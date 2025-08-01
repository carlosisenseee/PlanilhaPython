import re
from xml.sax.handler import property_interning_dict

import requests
import json
from openpyxl import load_workbook

url = "https://apiintegracao.milvus.com.br/api/chamado/listagem"

payload = {
    "filtro_body": {
        "tipo_ticket": "Novo Colaborador"
    }
}

headers = {
    "Content-Type": "application/json",
    "Authorization": "7X0CuWWQydCeiBWRpp1fTipDQjC3mWqw7pmUjEdMySCY0iB1KyRm1uWfp4l8BBoWrULb2JvlDskE9YLTd1D8v8x4QhSFlYS5QCknx"
}

response = requests.post(url, headers=headers, json=payload, timeout=30)
response.raise_for_status()

dados = response.json()

wb = load_workbook("Criação de Acessos.xlsx")
ws = wb["Planilha1"]

contador = 0

for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    valor = row[0]

    if isinstance(valor, (int, float)) and 1000 <= valor <= 9999:
        contador += 1

linha_atual = contador + 2

for chamado in dados.get("lista", []):
    ws.cell(row=linha_atual, column=1).value = chamado["codigo"]

    padrao = re.compile(
        r'^(?P<setor_lider>[\w\s]+\/[\w\s]+)\s+[–-]\s+(?P<nome>.+?)\s+[–-]\s+(?P<data>\d{2}/\d{2})\s+[–-]\s+(?P<local>.+?)\s*(?:-[–-]\s*Vaga\s+\d+)?$',
        re.UNICODE
    )

    print()
    match = padrao.match(chamado["assunto"])
    if match:
        ws.cell(row=linha_atual, column=2).value = match.group("setor_lider")
        ws.cell(row=linha_atual, column=3).value = match.group("nome")
        ws.cell(row=linha_atual, column=4).value = match.group("data")
        ws.cell(row=linha_atual, column=5).value = match.group("local")

    linha_atual += 1



wb.save("Criação de Acessos.xlsx")
print("Chamado foi")